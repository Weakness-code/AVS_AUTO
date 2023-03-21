from os import path, getenv
import datetime
import tkinter as tk
from openpyxl import Workbook, load_workbook
from porduct_creater import extractor
from tkinter import scrolledtext
from dataclasses import dataclass


price_list = 'pricelist.xlsx'


@dataclass
class Product:
    Id: str = ''
    DateBegin: str = ''
    DateEnd: str = ''
    ManagerName: str = getenv('ManagerName')
    ContactPhone: str = getenv("ContactPhone")
    Address: str = getenv("Address")
    Category: str = "Запчасти и аксессуары"
    Title: str = ''
    Description: str = ''
    Price: float = .0
    ImageUrls: str = ''
    VideoUrl: str = ''
    GoodsType: str = "Инструменты"
    AdType: str = "Товар приобретен на продажу"
    Condition: str = "Новое"
    Brand: str = "AVS"
    OEM: str = ''


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.pack()

        self.label_range = tk.Label(self, text="Промежуток = ", borderwidth=5)
        self.label_category = tk.Label(self, text="Категория = ", borderwidth=5)
        self.label_goodstype = tk.Label(self, text="Подкатегория = ", borderwidth=5)
        self.text_range = tk.Text(self, height=1, width=30)
        self.text_category = tk.Text(self, height=1, width=30)
        self.text_goodstype = tk.Text(self, height=1, width=30)
        self.scrollbox = scrolledtext.ScrolledText(self, width=40, height=18, undo=True)
        self.button = tk.Button(self, text="Создать", command=self.excel_creator)
        self.button_update = tk.Button(self, text="Обновить статус товаров", command=self.products_update)

        self.create_widgets()

    def create_widgets(self):
        self.label_range.grid(row=1, column=1, sticky=tk.NW)
        self.label_category.grid(row=2, column=1, sticky=tk.NW)
        self.label_goodstype.grid(row=3, column=1, sticky=tk.NW)
        self.text_range.grid(row=1, column=2)
        self.text_category.grid(row=2, column=2)
        self.text_goodstype.grid(row=3, column=2)
        self.scrollbox.grid(row=4, columnspan=3, pady=5)
        self.button.grid(row=5, column=2, pady=10, padx=10, sticky=tk.SE)
        self.button_update.grid(row=5, column=1, columnspan=2, sticky=tk.W)

    def excel_creator(self):
        try:
            excel_book = load_workbook('Products.xlsx')
            excel_sheet = excel_book.active
        except FileNotFoundError:
            excel_book = Workbook()
            excel_sheet = excel_book.active
            excel_sheet.append([i for i in Product.__annotations__.keys()])
        try:
            workbook = load_workbook(price_list)
        except FileNotFoundError:
            self.scrollbox.insert('end', f"Неверный формат файла или файл отсутсвует")
            return
        worksheet = workbook.active
        work_cells = worksheet[self.text_range.get('1.0', 'end')]

        count = 0
        created = [i.value for i in excel_sheet['A']]
        for cell in work_cells:
            if cell[6].value != None:
                count += 1
        for cell in work_cells:
            if cell[6].value in ['Ожидается', 'В наличии'] and cell[0].value not in created:
                product = extractor(cell, Product)
                product.Category = self.text_category.get('1.0', 'end')
                product.GoodsType = self.text_goodstype.get('1.0', 'end')
                excel_sheet.append([i for i in product.__dict__.values()])

                count -= 1
                self.scrollbox.insert('end', f"{product.Id} - parsed, remaining {count}\n")
                self.update()
                
        self.scrollbox.insert('end', f"\n-------DONE-------\n")
        excel_book.save('Products.xlsx')

    def products_update(self):
        try:
            product_book = load_workbook('Products.xlsx')
        except FileNotFoundError:
            self.scrollbox.insert('end', f"Файл отсутсвует")
            return
        product_sheet = product_book.active
        product_cells = product_sheet['A:S']
        try:
            price_book = load_workbook(price_list)
        except FileNotFoundError:
            self.scrollbox.insert('end', f"Неверный формат файла или файл отсутсвует")
            return
        price_sheet = price_book.active
        price_id = price_sheet['A']
        price_status = price_sheet['G']
        articul = price_sheet['B']
        
        time = datetime.datetime.now().astimezone().replace(microsecond=0).isoformat()
        for i in range(1, len(product_cells[0])):
            for j in range(len(price_id)):
                if product_cells[0][i].value == price_id[j].value:
                    product_cells[1][i].value = time if price_status[j].value == 'В наличии' else ''
                    product_cells[2][i].value = time if price_status[j].value != 'В наличии' else ''
                    product_cells[17][i].value = "AVS"
                    product_cells[18][i].value = articul[j].value
                    break
            self.update()
                
        self.scrollbox.insert('end', f"\n-------DONE-------\n")
        product_book.save('Products.xlsx')
    
    
def main():
    root = tk.Tk()
    root.title("Avs-Auto")
    root.geometry('400x450')

    app = Application(root)
    app.mainloop()


if __name__ == '__main__':
    main()
